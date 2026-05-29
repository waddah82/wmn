import frappe
from frappe import _
import traceback
import io
import base64
from openpyxl import load_workbook
import csv

@frappe.whitelist()
def parse_excel_file11(file_url, docname):
    """Parse Excel/CSV file and populate the table"""
    try:
        # التحقق من وجود الوثيقة
        if not docname:
            frappe.throw(_("Document name is required. Please save the document first."))
        
        # التحقق من أن الوثيقة موجودة في قاعدة البيانات
        if not frappe.db.exists("Stock Balance Import", docname):
            frappe.throw(_("Document {0} not found. Please save the document first.").format(docname))
        
        # تحميل الملف
        file_doc = frappe.get_doc("File", {"file_url": file_url})
        if not file_doc:
            frappe.throw(_("File not found"))
        
        file_path = file_doc.get_full_path()
        
        frappe.log_error(f"Reading file: {file_path} for doc: {docname}", "Excel Import")
        
        # قراءة البيانات
        data_rows = []
        headers = []
        
        if file_url.endswith('.csv'):
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                headers = next(reader)
                for row in reader:
                    if row and any(cell.strip() for cell in row):
                        data_rows.append(row)
        else:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and any(cell is not None and str(cell).strip() for cell in row):
                    data_rows.append(row)
        
        # البحث عن الأعمدة
        item_col = find_column(headers, ['item', 'product', 'code', 'barcode', 'sku'])
        qty_col = find_column(headers, ['qty', 'quantity', 'amount', 'count'])
        rate_col = find_column(headers, ['rate', 'price', 'valuation', 'cost', 'val'])
        expiry_col = find_column(headers, ['expiry', 'expiration', 'expire', 'date'])
        
        if item_col is None and len(headers) > 0:
            item_col = 0
        if qty_col is None and len(headers) > 1:
            qty_col = 1
        if rate_col is None and len(headers) > 2:
            rate_col = 2
        
        # معالجة البيانات
        items = []
        skipped_rows = 0
        
        for row_index, row in enumerate(data_rows, start=2):
            try:
                item_code = None
                if item_col is not None and item_col < len(row):
                    val = row[item_col]
                    if val is not None and str(val).strip():
                        item_code = str(val).strip()
                
                qty = 0
                if qty_col is not None and qty_col < len(row):
                    val = row[qty_col]
                    if val is not None:
                        try:
                            qty = float(val)
                        except:
                            qty = 0
                
                rate = 0
                if rate_col is not None and rate_col < len(row):
                    val = row[rate_col]
                    if val is not None:
                        try:
                            rate = float(val)
                        except:
                            rate = 0
                
                expiry = None
                if expiry_col is not None and expiry_col < len(row):
                    val = row[expiry_col]
                    if val is not None and str(val).strip():
                        try:
                            expiry = convert_to_date(val)
                        except:
                            expiry = None
                
                if not item_code or item_code == 'nan' or item_code == '':
                    skipped_rows += 1
                    continue
                
                if qty <= 0:
                    skipped_rows += 1
                    continue
                
                items.append({
                    "item_code": item_code,
                    "batch_qty": qty,
                    "valuation_rate": rate,
                    "expiry_date": expiry
                })
                
            except Exception as row_error:
                frappe.log_error(f"Row {row_index} error: {str(row_error)}", "Excel Import Row Error")
                skipped_rows += 1
                continue
        
        if not items:
            frappe.throw(_("No valid data found in the file. Please check the format."))
        
        # تحديث الوثيقة الموجودة
        doc = frappe.get_doc("Stock Balance Import", docname)
        doc.set('items_table', [])
        
        for item in items:
            doc.append('items_table', item)
        
        doc.save()
        frappe.db.commit()
        
        return {
            "success": True,
            "row_count": len(items),
            "skipped_rows": skipped_rows,
            "docname": docname,
            "message": f"Successfully imported {len(items)} rows. Skipped {skipped_rows} invalid rows."
        }
        
    except Exception as e:
        frappe.log_error(f"Excel Import Error: {str(e)}\n{traceback.format_exc()}", "Excel Import Error")
        return {
            "success": False,
            "error": str(e)
        }


def parse_excel_date(date_value):
    """تحويل التاريخ من Excel مع دعم القيم الفارغة"""
    from datetime import datetime
    
    if date_value is None:
        return None
    
    if isinstance(date_value, datetime):
        return date_value.date()
    
    if isinstance(date_value, str):
        date_str = date_value.strip()
        if not date_str or date_str.lower() in ['', 'nan', 'null', 'none', 'empty']:
            return None
        
        # تنسيقات التاريخ المدعومة
        formats = [
            '%Y-%m-%d',  # 2025-12-31
            '%d/%m/%Y',  # 31/12/2025
            '%m/%d/%Y',  # 12/31/2025
            '%d-%m-%Y',  # 31-12-2025
            '%m-%d-%Y',  # 12-31-2025
            '%Y%m%d',    # 20251231
            '%d.%m.%Y',  # 31.12.2025
            '%b %d, %Y', # Dec 31, 2025
            '%d %b %Y',  # 31 Dec 2025
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except:
                continue
        
        # إذا فشل كل شيء، سجل تحذير
        frappe.log_error(f"Could not parse date: {date_str}", "Excel Date Parse Error")
        return None
    
    return None

@frappe.whitelist()
def parse_excel_file(file_url, docname):
    """Parse Excel/CSV file with improved date handling"""
    try:
        # التحقق من وجود الوثيقة
        if not docname:
            frappe.throw(_("Document name is required. Please save the document first."))
        
        if not frappe.db.exists("Stock Balance Import", docname):
            frappe.throw(_("Document {0} not found. Please save the document first.").format(docname))
        
        # تحميل الملف
        file_doc = frappe.get_doc("File", {"file_url": file_url})
        if not file_doc:
            frappe.throw(_("File not found"))
        
        file_path = file_doc.get_full_path()
        
        frappe.log_error(f"Reading file: {file_path} for doc: {docname}", "Excel Import")
        
        # قراءة البيانات
        data_rows = []
        headers = []
        
        if file_url.endswith('.csv'):
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                headers = next(reader)
                for row in reader:
                    if row and any(cell.strip() for cell in row):
                        data_rows.append(row)
        else:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and any(cell is not None and str(cell).strip() for cell in row):
                    data_rows.append(row)
        
        # البحث عن الأعمدة
        item_col = find_column(headers, ['item', 'product', 'code', 'barcode', 'sku'])
        qty_col = find_column(headers, ['qty', 'quantity', 'amount', 'count'])
        rate_col = find_column(headers, ['rate', 'price', 'valuation', 'cost', 'val'])
        expiry_col = find_column(headers, ['expiry', 'expiration', 'expire', 'date'])
        
        if item_col is None and len(headers) > 0:
            item_col = 0
        if qty_col is None and len(headers) > 1:
            qty_col = 1
        if rate_col is None and len(headers) > 2:
            rate_col = 2
        
        # معالجة البيانات
        items = []
        skipped_rows = []
        row_errors = {}
        
        for row_index, row in enumerate(data_rows, start=2):
            errors = []
            
            try:
                # قراءة البيانات مع تحسين معالجة التاريخ
                item_code = None
                if item_col is not None and item_col < len(row):
                    val = row[item_col]
                    if val is not None and str(val).strip() and str(val).lower() not in ['nan', 'null', 'none']:
                        item_code = str(val).strip()
                
                qty = 0
                if qty_col is not None and qty_col < len(row):
                    val = row[qty_col]
                    if val is not None:
                        try:
                            qty = float(val)
                        except:
                            qty = 0
                
                rate = 0
                if rate_col is not None and rate_col < len(row):
                    val = row[rate_col]
                    if val is not None:
                        try:
                            rate = float(val)
                        except:
                            rate = 0
                
                # معالجة التاريخ المحسنة
                expiry = None
                if expiry_col is not None and expiry_col < len(row):
                    val = row[expiry_col]
                    if val is not None and str(val).strip():
                        # التحقق من القيم الفارغة
                        val_str = str(val).strip().lower()
                        if val_str not in ['', 'nan', 'null', 'none', 'empty', 'na']:
                            expiry = parse_excel_date(val)
                
                # التحقق من صحة البيانات
                if not item_code:
                    errors.append("Item Code is empty")
                
                if qty <= 0:
                    errors.append(f"Quantity must be greater than zero (current: {qty})")
                
                if rate <= 0:
                    errors.append(f"Valuation rate must be greater than zero (current: {rate})")
                
                # إذا كان هناك أخطاء، تخطي هذا الصف
                if errors:
                    skipped_rows.append(row_index)
                    row_errors[row_index] = errors
                    continue
                
                # إضافة للقائمة
                items.append({
                    "item_code": item_code,
                    "batch_qty": qty,
                    "valuation_rate": rate,
                    "expiry_date": expiry
                })
                
            except Exception as row_error:
                errors.append(f"Error processing row: {str(row_error)}")
                skipped_rows.append(row_index)
                row_errors[row_index] = errors
                continue
        
        if not items:
            error_message = "No valid data found in the file.\n\n"
            if row_errors:
                error_message += "Errors found:\n"
                for row_num, errs in row_errors.items():
                    error_message += f"Row {row_num}: {', '.join(errs)}\n"
            frappe.throw(_(error_message))
        
        # تحديث الوثيقة
        doc = frappe.get_doc("Stock Balance Import", docname)
        doc.set('items_table', [])
        
        for item in items:
            doc.append('items_table', item)
        
        doc.save()
        frappe.db.commit()
        
        # إعداد رسالة النتيجة
        message = f"Successfully imported {len(items)} rows."
        if skipped_rows:
            message += f"\n\n⚠️ Skipped {len(skipped_rows)} rows due to errors."
            message += f"\n\nErrors details:\n"
            for row_num, errs in row_errors.items():
                message += f"\nRow {row_num}: {', '.join(errs)}"
        
        return {
            "success": True,
            "row_count": len(items),
            "skipped_rows": len(skipped_rows),
            "skipped_row_numbers": skipped_rows,
            "row_errors": row_errors,
            "message": message
        }
        
    except Exception as e:
        frappe.log_error(f"Excel Import Error: {str(e)}\n{traceback.format_exc()}", "Excel Import Error")
        return {
            "success": False,
            "error": str(e)
        }




def find_column(headers, keywords):
    """Find column index by keywords"""
    for i, header in enumerate(headers):
        if header:
            header_lower = str(header).lower()
            for keyword in keywords:
                if keyword in header_lower:
                    return i
    return None

def convert_to_date(date_value):
    """Convert various date formats to date object"""
    from datetime import datetime
    
    if isinstance(date_value, datetime):
        return date_value.date()
    
    if isinstance(date_value, str):
        formats = [
            '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y',
            '%d-%m-%Y', '%m-%d-%Y', '%Y%m%d',
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(date_value.strip(), fmt).date()
            except:
                continue
    
    return None

@frappe.whitelist()
def get_sample_template():
    """Generate a sample Excel template for download"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.comments import Comment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Balance Template"
        
        # Add headers
        headers = ['Item Code', 'Quantity', 'Valuation Rate', 'Expiry Date']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add sample data
        sample_data = [
            ['3350900000011', 41, 48, '2029-05-01'],
            ['8809626560750', 69, 33.5, '2027-09-04'],
            ['8809809428327', 18, 35, '2027-07-01'],
        ]
        
        for row_idx, row_data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Add comments
        ws.cell(row=1, column=1).comment = Comment("Required: Item code, barcode, or SKU", "System")
        ws.cell(row=1, column=2).comment = Comment("Required: Quantity (must be greater than 0)", "System")
        ws.cell(row=1, column=3).comment = Comment("Required: Valuation rate/price", "System")
        ws.cell(row=1, column=4).comment = Comment("Optional: Expiry date (YYYY-MM-DD format)", "System")
        
        # Adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[chr(64 + col)].width = 20
        
        # Save file
        output = io.BytesIO()
        wb.save(output)
        excel_data = output.getvalue()
        encoded = base64.b64encode(excel_data).decode('utf-8')
        
        return {
            "success": True,
            "file_content": encoded,
            "filename": "stock_balance_template.xlsx"
        }
        
    except Exception as e:
        frappe.log_error(f"Template generation error: {str(e)}", "Excel Template Error")
        return {
            "success": False,
            "error": str(e)
        }